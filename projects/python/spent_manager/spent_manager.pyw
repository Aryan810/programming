import os
from tkinter import *
from datetime import datetime
import openpyxl as opx
from tkinter.messagebox import showwarning

FILENAME = "transactions.xlsx"
BG1 = "ghost white"
HEIGHT = 0
WIDTH = 0
FONT_1 = ("Roboto", 20)
FONT_2 = ("Roboto", 26)
FONT_3 = ("Roboto", 19)
FONT_4 = ("Roboto", 14)

'''
format of saving transactions


'''


def get_date():
    now = datetime.now()
    return now.strftime("%d %b, %Y --- %I:%M %p")


class SpentManager:

    def __init__(self, main_: Tk):
        self.data = {}
        self.n = 1
        self.row = 1
        self.main = main_
        self.total_spent_money = 0
        self.total_spent_money_var = StringVar(self.main, value="0")
        self.date_var = StringVar(self.main, value=get_date())
        self.amount_var = StringVar(self.main, value="0000")
        self.reason_var = StringVar(self.main)
        self.desc_var = StringVar(self.main)
        self.trans_history_lst = Listbox()

        self.reload()

        self.make_gui(self.main)
        self.edit_trans_lst()

    def edit_trans_lst(self):
        for i in self.data.keys():
            self.trans_history_lst.insert(END, f"{i} :) {self.data[i][0]} => Rs {self.data[i][3]}")
            if self.data[i][1] != None:
                self.trans_history_lst.insert(END, "\tReason: " + str(self.data[i][1]))
            if self.data[i][2] != None:
                self.trans_history_lst.insert(END, "\tDescription: " + str(self.data[i][2]))
            self.trans_history_lst.insert(END, "-"*100)

    def reload(self):
        if os.path.exists(FILENAME):
            wb = opx.load_workbook(FILENAME)
            sheet = wb.active
            self.data = {}

            '''
            data -> (serial no.) : [date, reason, description, amount]
            '''

            while (sheet.cell(self.row, 1).value != None):
                self.data[self.row] = [sheet.cell(self.row, 1).value,
                                       sheet.cell(self.row, 2).value,
                                       sheet.cell(self.row, 3).value,
                                       sheet.cell(self.row, 4).value]
                self.row += 1
            for value in self.data.values():
                self.total_spent_money += float(value[3])
            self.total_spent_money_var.set("Rs " + str(self.total_spent_money))
            wb.save(FILENAME)
        else:
            wb = opx.Workbook()
            wb.save(FILENAME)
            self.data = {}

    def add_transaction(self):
        trans_time = self.date_var.get()
        trans_reason = self.reason_var.get()
        trans_description = self.desc_var.get()

        try:
            trans_amount = float(self.amount_var.get())
            if trans_amount == 0.0:
                raise
            wb = opx.load_workbook(FILENAME)
            sheet = wb.active
            sheet.append([trans_time, trans_reason, trans_description, trans_amount])
            wb.save(FILENAME)
            self.reload()
            self.edit_trans_lst()
        except Exception as e:
            print(e)
            showwarning("Spent Manager - Warning", "Amount should be a proper non zero numerical value!")

    def make_gui(self, master):
        left_frame = Frame(master)

        add_trans_frame = Frame(left_frame)

        # This frame takes all entry windows
        date_frame = Frame(add_trans_frame)
        Label(date_frame, text="Date", font=FONT_1).pack(expand=True, fill=BOTH, padx=5, pady=10, side=LEFT)
        Entry(date_frame, textvariable=self.date_var, font=FONT_3).pack(expand=True, fill=BOTH, padx=5, pady=10, side=RIGHT)
        date_frame.pack(fill=BOTH, expand=True, side=TOP)

        amount_frame = Frame(add_trans_frame)
        Label(amount_frame, text="Amount", font=FONT_1).pack(expand=True, fill=BOTH, padx=5, pady=10, side=LEFT)
        Entry(amount_frame, textvariable=self.amount_var, font=FONT_3).pack(expand=True, fill=BOTH, padx=5, pady=10, side=RIGHT)
        amount_frame.pack(fill=BOTH, expand=True, side=TOP)

        reason_frame = Frame(add_trans_frame)
        Label(reason_frame, text="Reason", font=FONT_1).pack(expand=True, fill=BOTH, padx=5, pady=10, side=LEFT)
        Entry(reason_frame, textvariable=self.reason_var, font=FONT_3).pack(expand=True, fill=BOTH, padx=5, pady=10, side=RIGHT)
        reason_frame.pack(fill=BOTH, expand=True, side=TOP)

        desc_frame = Frame(add_trans_frame)
        Label(desc_frame, text="Description", font=FONT_1).pack(expand=True, fill=BOTH, padx=5, pady=10, side=LEFT)
        Entry(desc_frame, textvariable=self.desc_var, font=FONT_3).pack(expand=True, fill=BOTH, padx=5, pady=10, side=RIGHT)
        desc_frame.pack(fill=BOTH, expand=True, side=TOP)

        # This frame takes a button for adding transition
        button_frame = Frame(add_trans_frame)
        Button(button_frame, text="Add Transition", font=FONT_1, command=self.add_transaction).pack(expand=True, fill=BOTH, padx=10, pady=10)
        button_frame.pack(fill=BOTH, expand=True, side=BOTTOM)

        add_trans_frame.pack(fill=BOTH, expand=True, side=TOP)

        # Shows total spent money data
        border_color2 = Frame(left_frame, background="black")
        total_money_frame = Frame(border_color2, bg="ghost white")
        Label(total_money_frame, text="Total Money Spent: ", font=FONT_1, bg="ghost white").pack(pady=10, expand=True, fill=BOTH, side=LEFT)
        Label(total_money_frame, textvariable=self.total_spent_money_var, font=FONT_2, bg="ghost white").pack(pady=10, expand=True, fill=BOTH, side=RIGHT)

        total_money_frame.pack(fill=BOTH, side=BOTTOM, expand=True, padx=3, pady=3)

        left_frame.pack(fill=BOTH, expand=True, side=LEFT)
        border_color2.pack(fill=BOTH, expand=True, padx=3, pady=3)

        # --------- here left frame ends ------------

        list_trans_frame = Frame(master, bg="green", width=WIDTH/2.2, height=HEIGHT)
        list_and_label_frame = Frame(list_trans_frame)
        Label(list_and_label_frame, text="Transaction History", font=FONT_3).pack(padx=20, pady=10, side=TOP)
        border_color1 = Frame(list_and_label_frame, bg="black")
        self.trans_history_lst = Listbox(border_color1, relief=FLAT, font=FONT_4, bg=BG1)
        self.trans_history_lst.pack(pady=3, padx=3, expand=True, side=BOTTOM, fill=BOTH)
        border_color1.pack(fill=BOTH, expand=True, pady=3)
        scroll_bar = Scrollbar(list_trans_frame)
        scroll_bar.pack(side=RIGHT, fill=BOTH)
        self.trans_history_lst.config(yscrollcommand=scroll_bar.set)
        scroll_bar.config(command=self.trans_history_lst.yview)
        list_and_label_frame.pack(side=LEFT, fill=BOTH, expand=True)
        list_trans_frame.pack(fill=BOTH, expand=True, side=RIGHT)


if __name__ == "__main__":
    main = Tk()
    WIDTH = int(main.winfo_screenwidth()/1.9)
    HEIGHT = int(main.winfo_screenheight()/2.2)
    main.title("Spent Manager")
    main.minsize(WIDTH, HEIGHT)
    # main.maxsize(WIDTH, HEIGHT)

    sm = SpentManager(main)

    main.geometry(f"{WIDTH}x{HEIGHT}")

    main.mainloop()
